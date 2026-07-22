from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE_DIR = REPO_ROOT / "shuju"
DEFAULT_OUTPUT = REPO_ROOT / "backend" / "data" / "incidents.json"

DISTRICT_FROM_FILENAME = {
    "永定区": "永定区",
    "武陵源区": "武陵源区",
    "慈利县": "慈利县",
    "桑植县": "桑植县",
}

DISTRICT_CENTERS = {
    "永定区": (29.1612, 110.4797),
    "武陵源区": (29.3470, 110.5488),
    "慈利县": (29.3998, 110.9045),
    "桑植县": (29.5827, 110.1824),
}

DISTRICT_SHELTER_IDS = {
    "永定区": "zjj-safe-tianmenshan",
    "武陵源区": "zjj-safe-wly-weather",
    "慈利县": "sh-dxg-center",
    "桑植县": "zjj-safe-sz-yuanzicun",
}

LOCATION_ANCHORS = {
    "永定区": [
        (("天门山",), (29.1167, 110.4759)),
        (("老道湾",), (29.1584, 110.6253)),
        (("罗水乡", "大明村"), (29.2105, 110.3568)),
        (("新桥镇", "远大村"), (29.1584, 110.6253)),
    ],
    "武陵源区": [
        (("黄龙洞",), (29.3610, 110.6269)),
        (("金鞭溪",), (29.3472, 110.5587)),
        (("森林公园", "老磨湾"), (29.3342, 110.4461)),
        (("索溪峪",), (29.3457, 110.5488)),
        (("天子山",), (29.3338, 110.5145)),
    ],
    "慈利县": [
        (("大峡谷",), (29.3939, 110.6938)),
        (("长潭河", "金慈街道"), (29.4019, 111.1174)),
    ],
    "桑植县": [
        (("院子村", "上洞街"), (29.5827, 110.1824)),
        (("八大公山",), (29.7030, 110.1260)),
        (("澧源镇",), (29.4010, 110.1640)),
    ],
}

START_KEYS = ("发生起始时间", "起始年月日")
END_KEYS = ("发生结束时间", "结束时间", "结束年月日")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import shuju Excel files into backend/data/incidents.json")
    parser.add_argument("--source-dir", default=str(DEFAULT_SOURCE_DIR), help="Directory containing the shuju Excel files")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Target incidents.json path")
    return parser.parse_args()


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return datetime.combine(value, time.min).isoformat(sep=" ", timespec="seconds")
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def pick(row: dict, *keys: str) -> str:
    for key in keys:
        value = clean_text(row.get(key))
        if value:
            return value
    return ""


def join_non_empty(*values: str) -> str:
    return "；".join([clean_text(value) for value in values if clean_text(value)])


def parse_datetime(value, end_of_day: bool = False) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        base = datetime.combine(value, time.min)
        return base.replace(hour=23, minute=59, second=59) if end_of_day else base

    text = clean_text(value)
    if not text:
        return None
    normalized = text.replace("年", "-").replace("月", "-").replace("日", "").replace("/", "-")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(normalized, fmt)
            if fmt == "%Y-%m-%d" and end_of_day:
                parsed = parsed.replace(hour=23, minute=59, second=59)
            return parsed
        except ValueError:
            continue
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", normalized)
    if match:
        parsed = datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        if end_of_day:
            parsed = parsed.replace(hour=23, minute=59, second=59)
        return parsed
    return None


def first_header_row(rows: list[tuple]) -> int:
    for index, row in enumerate(rows[:3]):
        if clean_text(row[0]) == "序号":
            return index
    return 0


def district_from_filename(path: Path) -> str:
    for key, district in DISTRICT_FROM_FILENAME.items():
        if key in path.stem:
            return district
    raise ValueError(f"Cannot infer district from filename: {path.name}")


def stable_offset(seed: str, scale: float) -> tuple[float, float]:
    digest = hashlib.md5(seed.encode("utf-8")).hexdigest()
    lat_raw = int(digest[:8], 16) / 0xFFFFFFFF
    lng_raw = int(digest[8:16], 16) / 0xFFFFFFFF
    return ((lat_raw - 0.5) * scale, (lng_raw - 0.5) * scale)


def infer_coordinate(district: str, title: str, location: str, seed_index: int) -> tuple[float, float]:
    text = f"{title} {location}"
    base = DISTRICT_CENTERS[district]
    scale = 0.05
    for keywords, point in LOCATION_ANCHORS.get(district, []):
        if all(keyword in text for keyword in keywords):
            base = point
            scale = 0.018
            break
        if any(keyword in text for keyword in keywords):
            base = point
            scale = 0.026
            break

    lat_offset, lng_offset = stable_offset(f"{district}:{title}:{seed_index}", scale)
    return round(base[0] + lat_offset, 6), round(base[1] + lng_offset, 6)


def extract_primary_place(district: str, location: str) -> str:
    text = clean_text(location)
    if not text:
        return district

    text = text.replace(f"张家界市{district}", "").strip("，,；; ")
    for marker in ("核心点位：", "重灾区："):
        if marker in text:
            text = text.split(marker, 1)[1]
            break

    tokens = [token.strip() for token in re.split(r"[；;，,、]", text) if token.strip()]
    for token in tokens:
        normalized = token.replace("全域", "").strip()
        if normalized:
            return normalized
    return district


def infer_community(primary_place: str) -> str | None:
    if not primary_place:
        return None
    if any(keyword in primary_place for keyword in ("村", "社区", "街道", "乡", "镇", "景区")):
        return primary_place
    return None


def classify_type(title: str, disaster_type: str, location: str, people: str) -> str:
    hazard_text = f"{title} {disaster_type} {location}"
    people_text = clean_text(people)
    if any(keyword in hazard_text for keyword in ("干旱", "伏旱", "风雹", "冰雹", "大风")):
        return "other"
    if any(keyword in hazard_text for keyword in ("冰冻", "雪灾", "道路结冰")):
        return "road"
    if any(keyword in hazard_text for keyword in ("洪涝", "山洪", "暴雨", "积水", "涨水", "内涝")):
        return "flood"
    if any(keyword in hazard_text for keyword in ("滑坡", "泥石流", "崩塌", "落石")):
        return "landslide"
    if any(keyword in hazard_text for keyword in ("道路", "桥梁", "中断")):
        return "road"
    if any(keyword in people_text for keyword in ("SOS", "求助", "被困")):
        return "sos"
    if any(keyword in hazard_text for keyword in ("医疗", "救护")):
        return "medical"
    if any(keyword in hazard_text for keyword in ("安置点", "避险转移")):
        return "shelter"
    return "other"


def extract_number(text: str, keyword: str) -> int:
    patterns = [
        rf"{re.escape(keyword)}[^\d]{{0,8}}(\d+)",
        rf"(\d+)\s*(?:人|户|间|处|个|名|台|公里|km|m|亩|万元|亿元)?[^\d]{{0,4}}{re.escape(keyword)}",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return int(match.group(1))
    return 0


def parse_amount_yuan(text: str) -> float:
    source = clean_text(text)
    if not source:
        return 0.0
    match = re.search(r"(\d+(?:\.\d+)?)\s*(亿元|万元|元)", source)
    if not match:
        return 0.0
    amount = float(match.group(1))
    unit = match.group(2)
    if unit == "亿元":
        return amount * 100_000_000
    if unit == "万元":
        return amount * 10_000
    return amount


def infer_severity(title: str, people: str, direct_loss: str, incident_type: str) -> str:
    text = f"{title} {people}"
    deaths = max(extract_number(text, "死亡"), extract_number(text, "遇难"))
    trapped = extract_number(text, "被困")
    injured = extract_number(text, "受伤")
    transferred = max(extract_number(text, "转移"), extract_number(text, "安置"))
    direct_loss_yuan = parse_amount_yuan(direct_loss)

    if deaths > 0 or trapped >= 50 or transferred >= 1000 or direct_loss_yuan >= 100_000_000:
        return "critical"
    if "特大" in title or injured > 0 or transferred >= 200 or direct_loss_yuan >= 20_000_000:
        return "high"
    if incident_type in {"flood", "landslide", "road"}:
        return "medium"
    return "low"


def infer_status(start_dt: datetime | None, end_dt: datetime | None) -> str:
    now = datetime.now()
    if start_dt and start_dt > now:
        return "pending"
    if end_dt and end_dt < now:
        return "resolved"
    if start_dt and start_dt <= now and (end_dt is None or end_dt >= now):
        return "responding"
    return "resolved"


def build_workflow_steps(
    warning: str,
    rescue: str,
    supplies: str,
    shelter_setup: str,
    aftermath: str,
) -> list[str]:
    items = [
        ("预警发布", warning),
        ("救援处置", rescue),
        ("物资调拨", supplies),
        ("安置设置", shelter_setup),
        ("善后复盘", aftermath),
    ]
    return [f"{label}：{content}" for label, content in items if content]


def truncate(text: str, limit: int) -> str:
    source = clean_text(text)
    if len(source) <= limit:
        return source
    return f"{source[: limit - 1]}…"


def normalize_row(row: dict, district: str, seed_index: int) -> dict:
    title = pick(row, "灾害名称")
    location = pick(row, "发生地点")
    disaster_type = pick(row, "灾害类型")
    weather = pick(row, "受灾实况-气象数据", "气象数据", "气象实测数据")
    housing = join_non_empty(
        pick(row, "受灾实况-房屋情况", "房屋受灾情况"),
        f"房屋倒塌{pick(row, '房屋倒塌户数')}" if pick(row, "房屋倒塌户数") else "",
        f"房屋受损{pick(row, '房屋受损数量')}" if pick(row, "房屋受损数量") else "",
    )
    people = join_non_empty(
        pick(row, "受灾实况-人员情况", "人员受灾情况"),
        f"人员伤亡{pick(row, '人员伤亡人数')}" if pick(row, "人员伤亡人数") else "",
        f"紧急转移安置{pick(row, '紧急转移安置人数')}" if pick(row, "紧急转移安置人数") else "",
        f"被困人员{pick(row, '被困人员数量')}" if pick(row, "被困人员数量") else "",
    )
    property_loss = join_non_empty(
        pick(row, "受灾实况-财产情况", "财产受灾情况"),
        pick(row, "农田受灾/绝收面积"),
    )
    infrastructure = join_non_empty(
        pick(row, "受灾实况-基础设施情况", "基础设施受灾情况"),
        pick(row, "道路损毁里程"),
        pick(row, "水利/电力/通讯设施损毁"),
        pick(row, "河堤/水库/景区受损情况"),
    )
    warning = join_non_empty(
        pick(row, "处置救援-预警发布情况", "预警发布时间", "预警发布时间与等级"),
        pick(row, "预警等级"),
    )
    rescue = pick(row, "处置救援-转移与救援队伍情况", "转移避险与救援出动", "救援处置与队伍出动")
    supplies = pick(row, "处置救援-抢险物资调拨", "抢险物资调拨")
    shelter_setup = pick(row, "处置救援-临时安置点设置", "临时安置点设置")
    aftermath = join_non_empty(
        pick(row, "善后与重建", "善后与复盘-灾后重建"),
        pick(row, "善后与复盘-救灾补助发放"),
        pick(row, "善后与复盘-隐患治理"),
        pick(row, "善后与复盘-经验教训"),
        pick(row, "损失数据与善后复盘"),
    )
    direct_loss = pick(row, "直接经济损失", "损失数据-直接经济损失")
    indirect_loss = pick(row, "间接损失", "损失数据-间接损失")
    insurance = pick(row, "保险赔付情况", "损失数据-保险赔付情况")

    start_dt = parse_datetime(row.get("发生起始时间") or row.get("起始年月日"))
    end_dt = parse_datetime(row.get("发生结束时间") or row.get("结束时间") or row.get("结束年月日"), end_of_day=True)
    primary_place = extract_primary_place(district, location)
    lat, lng = infer_coordinate(district, title, location, seed_index)
    incident_type = classify_type(title, disaster_type, location, people)
    severity = infer_severity(title, people, direct_loss, incident_type)
    status = infer_status(start_dt, end_dt)

    impact_summary = first_non_empty(
        people,
        property_loss,
        infrastructure,
        weather,
    )
    description = f"{title}；地点：{truncate(primary_place or district, 32)}；类型：{truncate(disaster_type or incident_type, 32)}"
    if impact_summary:
        description = f"{description}；影响：{truncate(impact_summary, 58)}"

    source_date = start_dt.date().isoformat() if start_dt else ""
    created_at = (start_dt or datetime.now()).isoformat(timespec="seconds")
    resolved_at = end_dt.isoformat(timespec="seconds") if end_dt else None

    shelter_id = DISTRICT_SHELTER_IDS[district]
    if district == "慈利县" and any(keyword in location for keyword in ("长潭河", "金慈街道")):
        shelter_id = "zjj-safe-cl-changtanhe"
    elif district == "武陵源区" and "黄龙洞" in location:
        shelter_id = "sh-hld"
    elif district == "武陵源区" and any(keyword in location for keyword in ("索溪峪", "天子山", "中湖")):
        shelter_id = "zjj-safe-wly-weather"

    return {
        "type": incident_type,
        "description": description,
        "lat": lat,
        "lng": lng,
        "district": district,
        "community": infer_community(primary_place),
        "scenic_area": primary_place or district,
        "severity": severity,
        "status": status,
        "reporter_role": "county_admin",
        "reporter_id": 1,
        "nearest_shelter_id": shelter_id,
        "source_title": title,
        "source_org": "张家界区县灾害事件台账",
        "source_url": "",
        "source_date": source_date,
        "workflow_steps": build_workflow_steps(warning, rescue, supplies, shelter_setup, aftermath),
        "created_at": created_at,
        "resolved_at": resolved_at,
        "is_demo": True,
        "meta": {
            "location": location,
            "weather": weather,
            "housing": housing,
            "people": people,
            "property_loss": property_loss,
            "infrastructure": infrastructure,
            "direct_loss": direct_loss,
            "indirect_loss": indirect_loss,
            "insurance": insurance,
        },
    }


def first_non_empty(*values: str) -> str:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return ""


def load_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = first_header_row(rows)
    headers = [clean_text(cell) for cell in rows[header_index]]
    items: list[dict] = []
    for raw_row in rows[header_index + 1 :]:
        if not any(cell not in (None, "") for cell in raw_row):
            continue
        row = {headers[index]: raw_row[index] for index in range(min(len(headers), len(raw_row)))}
        if not pick(row, "灾害名称"):
            continue
        items.append(row)
    return items


def build_incidents(source_dir: Path) -> list[dict]:
    incidents: list[dict] = []
    files = sorted(source_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No xlsx files found under {source_dir}")

    seed_index = 0
    for path in files:
        district = district_from_filename(path)
        for row in load_rows(path):
            incidents.append(normalize_row(row, district, seed_index))
            seed_index += 1

    incidents.sort(key=lambda item: item.get("created_at", ""), reverse=True)
    return incidents


def main() -> None:
    args = parse_args()
    source_dir = Path(args.source_dir).resolve()
    output_path = Path(args.output).resolve()
    incidents = build_incidents(source_dir)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(incidents, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Imported {len(incidents)} incidents into {output_path}")


if __name__ == "__main__":
    main()
